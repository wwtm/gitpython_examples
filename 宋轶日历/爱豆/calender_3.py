from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import openpyxl
import calendar

# 设置第一天为星期天
calendar.setfirstweekday(firstweekday=6)
# 创建一个工作䈬
wb = openpyxl.Workbook()

# 遍历12个月
for i in range(1, 13):
    # 添加工作表
    sheet = wb.create_sheet(index=0, title=str(i) + '月')
    # 获取具体日期时间
    for j in range(len(calendar.monthcalendar(2019, i))):
        for k in range(len(calendar.monthcalendar(2019, i)[j])):
            value = calendar.monthcalendar(2019, i)[j][k]
            # 将0值变为空值
            if value == 0:
                value = ''
                sheet.cell(row=j + 9, column=k + 1).value = value
            else:
                sheet.cell(row=j + 9, column=k + 1).value = value
                # 设置字体
                sheet.cell(row=j + 9, column=k + 1).font = Font(u'微软雅黑', size=11)
    # 单元格文字设置,右对齐,垂直居中
    align = Alignment(horizontal='right', vertical='center')
    # 单元格填充色属性设置
    fill = PatternFill("solid", fgColor="B9EBF7")
    # 对单元格进行颜色填充
    for k1 in range(1, 100):
        for k2 in range(1, 100):
            sheet.cell(row=k1, column=k2).fill = fill
    # 添加星期几信息行
    days = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
    num = 0
    for k3 in range(1, 8):
        sheet.cell(row=8, column=k3).value = days[num]
        sheet.cell(row=8, column=k3).alignment = align
        sheet.cell(row=8, column=k3).font = Font(u'微软雅黑', size=11)
        # 设置列宽12
        c_char = get_column_letter(k3)
        sheet.column_dimensions[get_column_letter(k3)].width = 12
        num += 1
    # 设置行高30
    for k4 in range(8, 14):
        sheet.row_dimensions[k4].height = 30
    # 合并单元格
    sheet.merge_cells('I1:P20')
    # 添加图片
    img = Image('huge_2.jpg')
    sheet.add_image(img, 'I1')

    # 添加年份及月份
    sheet.cell(row=3, column=1).value = '2019年'
    sheet.cell(row=4, column=1).value = str(i) + '月'
    # 设置年份及月份文本属性
    sheet.cell(row=3, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    sheet.cell(row=4, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    sheet.cell(row=3, column=1).alignment = align
    sheet.cell(row=4, column=1).alignment = align
# 保存文档
wb.save('爱豆日历.xlsx')